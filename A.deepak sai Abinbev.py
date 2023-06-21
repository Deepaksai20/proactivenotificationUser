import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from smtplib import SMTP_SSL
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


linkedin_username = "chnikki2003@gmail.com"
linkedin_password = "Nikki2003*"


recipient_email = "deepaksai93900@gmail.com"
sender_email = "ebrahimgani123@gmail.com"
sender_password = "fvczbouiykfldfxu"

chromedriver_path = "C:/Users/deepa/OneDrive/Desktop/ABinBev/chromedriver.exe"
excel_file_path = "C:/Users/deepa/OneDrive/Desktop/ABinBev/Abinbev.xlsx"


chrome_options = Options()
#chrome_options.add_argument("--headless")  # Run Chrome in headless mode


driver = webdriver.Chrome(service=Service(chromedriver_path), options=chrome_options)


def login_to_linkedin():
    driver.get("https://www.linkedin.com/login")
    time.sleep(2)

    username_input = driver.find_element(By.ID, "username")
    password_input = driver.find_element(By.ID, "password")

    username_input.send_keys(linkedin_username)
    password_input.send_keys(linkedin_password)
    password_input.send_keys(Keys.ENTER)
    time.sleep(5)


# Function to get the number of unread messages
def get_unread_messages():
    driver.get("https://www.linkedin.com/messaging/")
    time.sleep(3)

    wait = WebDriverWait(driver, 10)
    unread_messages_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#global-nav > div > nav > ul > li:nth-child(4) > a > div > div > li-icon > svg")))
    unread_messages_text = unread_messages_element.get_attribute("innerText")

    unread_messages = int(unread_messages_text) if unread_messages_text else 0
    return unread_messages



# Function to get the number of unread notifications
def get_unread_notifications():
    driver.get("https://www.linkedin.com/notifications/")
    time.sleep(3)

    wait = WebDriverWait(driver, 10)
    unread_notifications_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#global-nav > div > nav > ul > li:nth-child(5) > a > div > div > li-icon > svg")))
    unread_notifications_text = unread_notifications_element.get_attribute("innerText")

    unread_notifications = int(unread_notifications_text) if unread_notifications_text else 0
    return unread_notifications



# Function to compare current data with previous occurrence data
def compare_data(current_data):
    previous_data = ""
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        previous_data = sheet.cell(row=sheet.max_row, column=1).value
    except:
        pass

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=sheet.max_row + 1, column=1).value = current_data
    workbook.save(excel_file_path)

    if previous_data:
        return f"Comparison with previous occurrence: {current_data} vs {previous_data}"
    else:
        return "No previous occurrence data available"

# Function to send email notification
def send_email_notification(num_unread_messages, num_unread_notifications, comparison_result):
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = recipient_email
    message["Subject"] = "LinkedIn Unread Notifications Update"

    body = f"Number of Unread Messages: {num_unread_messages}\nNumber of Unread Notifications: {num_unread_notifications}\n\n{comparison_result}"
    message.attach(MIMEText(body, "plain"))

    with SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender_email, sender_password)
        server.send_message(message)

# Execute the script
try:
    login_to_linkedin()
    unread_messages = get_unread_messages()
    unread_notifications = get_unread_notifications()
    comparison_result = compare_data(f"Messages: {unread_messages}, Notifications: {unread_notifications}")
    send_email_notification(unread_messages, unread_notifications, comparison_result)
except Exception as e:
    print(f"An error occurred: {str(e)}")
finally:
    driver.quit()
